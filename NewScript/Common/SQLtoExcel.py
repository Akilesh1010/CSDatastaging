import openpyxl
import pandas as pd

#filepath = 'NewScript/TestData/Sql_CsMarketDataQuery.xlsx'
# database connection in testsuite

# writing data into excel
#writer = pd.excelWriter(filepath, engine='xlxswriter')
# create a dataframe
#dataframe = pd.read_sql()

wk = openpyxl.load_workbook("NewScript/TestData/AllCsMarketDataTestSedols.xlsx")


def fetch_number_of_rows_sql(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row


def fetch_cell_data_sql(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value