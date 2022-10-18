from openpyxl import load_workbook
import requests

import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
import warnings

wb = load_workbook('C://cs_datastaging/test_data/Instrument_TradeCondition_testdata.xlsx')  # Work Book
ws = wb['Sheet1']  # Work Sheet
ws1 = wb['Sheet2']  # Work Sheet
Service = ws['A']  # Column
Symbol = ws['B']
GQL_Query = ws['C']
query =  GQL_Query
GQL_Query_null = ws1['A']

Service_list = [Service[x].value for x in range (1,(len(Service)))]
Symbol_list = [Symbol[x].value for x in range (1,(len(Symbol)))]
GQL_Query_list = [GQL_Query[x].value for x in range (1,(len(GQL_Query)))]
GQL_Query_null_list = [GQL_Query_null[x].value for x in range (1,(len(GQL_Query_null)))]
print(Service_list)
print(Symbol_list)
print(GQL_Query_list)
print(GQL_Query_null_list)

gpql_json_data1=[]
webservice_Tra_pri_value=[]
webservice_Tra_vol_value=[]
webservice_vol_value=[]
webservice_Day_High_value=[]
webservice_Day_Low_value=[]
webservice_Open_value=[]
webservice_Pre_clos_value=[]
webservice_sedol_value=[]


url = 'http://datastaging1-uat.charles-stanley.co.uk/graphql/'
for l in  GQL_Query_null_list:
    a = requests.post(url, json={'query': l}, timeout=20)
    gpql_json_data1.append(json.loads(a.text))
gpql_json_data=[]
for i,j,k in zip(GQL_Query_list,Service_list,Symbol_list):
    r = requests.post(url, json={'query': i} , timeout=20)
     ##time.sleep(5)
    response_status_code = r.status_code
    gpql_json_data.append(json.loads(r.text))



    warnings.filterwarnings("ignore", category=DeprecationWarning)
    driver = webdriver.Chrome(executable_path="C:\\cs_datastaging\Browsers\chromedriver.exe")
    driver.get("http://cstris01.charles-stanley.co.uk/ids-ws/item.jsp")
    driver.maximize_window()
    driver.implicitly_wait(10)
    driver.switch_to.frame(driver.find_element(By.NAME, 'input'))
    driver.find_element(By.XPATH, "//*[@id='service']").clear()
    driver.find_element(By.XPATH, "//*[@id='service']").send_keys(j)
    driver.find_element(By.XPATH, "//*[@id='symbol']").send_keys(k)
    driver.find_element(By.XPATH, "//*[@id='submit']").click()
    driver.switch_to.default_content()
    ##time.sleep(5)
    driver.switch_to.frame(driver.find_element(By.NAME, 'content'))
    try:
        webservice_Tra_pri_value1 = driver.find_element(By.XPATH, "//*/tr[6]/td[2]").text
        webservice_Tra_vol_value1 = driver.find_element(By.XPATH, "//*/tr[57]/td[2]").text
        webservice_vol_value1 =     driver.find_element(By.XPATH, "//*/tr[26]/td[2]").text
        webservice_Day_High_value1 = driver.find_element(By.XPATH, "//*/tr[12]/td[2]").text
        webservice_Day_Low_value1 = driver.find_element(By.XPATH, "//*/tr[13]/td[2]").text
        webservice_Open_value1 = driver.find_element(By.XPATH, "//*/tr[18]/td[2]").text
        webservice_Pre_clos_value1 = driver.find_element(By.XPATH, "//*/tr[19]/td[2]").text
        webservice_sedol_value1 = driver.find_element(By.XPATH, "//*/tr[157]/td[2]").text

    except:
        webservice_Tra_pri_value.append('None')
        webservice_Tra_vol_value.append('None')
        webservice_vol_value.append('None')
        webservice_Day_High_value.append('None')
        webservice_Day_Low_value.append('None')
        webservice_Open_value.append('None')
        webservice_Pre_clos_value.append('None')
        webservice_sedol_value.append('None')


        continue
    else:
        webservice_Tra_pri_value.append(webservice_Tra_pri_value1)
        webservice_Tra_vol_value.append(webservice_Tra_vol_value1)
        webservice_vol_value.append(webservice_vol_value1)
        webservice_Day_High_value.append(webservice_Day_High_value1)
        webservice_Day_Low_value.append(webservice_Day_Low_value1)
        webservice_Open_value.append(webservice_Open_value1)
        webservice_Pre_clos_value.append(webservice_Pre_clos_value1)
        webservice_sedol_value.append(webservice_sedol_value1)





    ##print(gpql_json_data)
    ##for j,k in zip(Service_list,Symbol_list):

driver.close()



print(gpql_json_data)
print(gpql_json_data1)
print(webservice_Tra_pri_value)
print(webservice_Tra_vol_value)
print(webservice_vol_value)
print(webservice_Day_High_value)
print(webservice_Day_Low_value)
print(webservice_Open_value)
print(webservice_Pre_clos_value)
print(webservice_sedol_value)







