import openpyxl
wk= openpyxl.load_workbook("C://cs_datastaging/test_data/Instrumet_testdata.xlsx")

def fetch_number_of_rows(Sheetname):
    sh= wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname,row,cell):
    sh= wk[Sheetname]
    cell= sh.cell(int(row),int(cell))
    return cell.value