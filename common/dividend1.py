from openpyxl import load_workbook
import requests

import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
import warnings

wb = load_workbook('C://cs_datastaging/test_data/Instrumet_dvidend_testdata.xlsx')  # Work Book
ws = wb['Sheet1']  # Work Sheet
ws1 = wb['Sheet2']
Service = ws['A']  # Column
Symbol = ws['B']
GQL_Query = ws['C']
GQLDB_Query= ws1['A']
DB_Query = ws1['B']
query =  GQL_Query
query1= GQLDB_Query

Service_list = [Service[x].value for x in range (2,(len(Service)))]
Symbol_list = [Symbol[x].value for x in range (2,(len(Symbol)))]
GQL_Query_list = [GQL_Query[x].value for x in range (2,(len(GQL_Query)))]
GQLDB_Query_list= [GQLDB_Query[x].value for x in range (2,(len(GQLDB_Query)))]
sql_query_list = [DB_Query[x].value for x in range (2,(len(DB_Query)))]
print(Service_list)
print(Symbol_list)
print(GQL_Query_list)
print(GQLDB_Query_list)
print(sql_query_list)
gpql_json_data=[]
gpql_json_data2=[]
webservice_YIELD_value=[]
webservice_DIVIDEND_value=[]
webservice_EXDIVDATE_value=[]
webservice_DIVPAYDATE_value=[]
url = 'http://datastaging1-sit.charles-stanley.co.uk/graphql/'
for l in  GQLDB_Query_list:
    a = requests.post(url, json={'query': l}, timeout=20)
    gpql_json_data2.append(json.loads(a.text))


for i,j,k in zip(GQL_Query_list,Service_list,Symbol_list):
    r = requests.post(url, json={'query': i} , timeout=20)
    ##time.sleep(5)
    response_status_code = r.status_code
    gpql_json_data.append(json.loads(r.text))

    warnings.filterwarnings("ignore", category=DeprecationWarning)
    driver = webdriver.Chrome(executable_path="C:\\cs_datastaging\Browsers\chromedriver.exe")
    driver.get("http://cstris01.charles-stanley.co.uk/ids-ws/item.jsp")
    driver.maximize_window()
    driver.implicitly_wait(10)
    driver.switch_to.frame(driver.find_element(By.NAME, 'input'))
    driver.find_element(By.XPATH, "//*[@id='service']").clear()
    driver.find_element(By.XPATH, "//*[@id='service']").send_keys(j)
    driver.find_element(By.XPATH, "//*[@id='symbol']").send_keys(k)
    driver.find_element(By.XPATH, "//*[@id='submit']").click()
    driver.switch_to.default_content()
    ##time.sleep(5)
    driver.switch_to.frame(driver.find_element(By.NAME, 'content'))
    try:
        webservice_YIELD_value1 = driver.find_element(By.XPATH, "//*/tr[28]/td[2]").text
        webservice_DIVIDEND_value1 = driver.find_element(By.XPATH, "//*/tr[39]/td[2]").text
        webservice_EXDIVDATE_value1 = driver.find_element(By.XPATH, "//*/tr[31]/td[2]").text
        webservice_DIVPAYDATE_value1 = driver.find_element(By.XPATH, "//*/tr[30]/td[2]").text
    except:
        webservice_YIELD_value.append('None')
        webservice_DIVIDEND_value.append('None')
        webservice_EXDIVDATE_value.append('None')
        webservice_DIVPAYDATE_value.append('None')

        continue
    else:
        webservice_YIELD_value.append(webservice_YIELD_value1)
        webservice_DIVIDEND_value.append(webservice_DIVIDEND_value1)
        webservice_EXDIVDATE_value.append(webservice_EXDIVDATE_value1)
        webservice_DIVPAYDATE_value.append(webservice_DIVPAYDATE_value1)






driver.close()



print(gpql_json_data)
print(gpql_json_data2)
print(webservice_YIELD_value)
print(webservice_DIVIDEND_value)
print(webservice_EXDIVDATE_value)
print(webservice_DIVPAYDATE_value)



