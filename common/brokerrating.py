from jsonpath_rw import parse
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import hashlib
#from jsonpath_ng import json
import requests
import json
import xmltodict



wb = load_workbook("C://cs_datastaging/test_data/Instrumet_brokerrating_testdata.xlsx")
ws = wb['Sheet1']
GQL_Query = ws['A']
num = 100
GQL_Query_list = [GQL_Query[x].value for x in range (1,(len(GQL_Query)))]
current_date = datetime.today().strftime('%Y-%m-%d')
userid = "charles_stanley_test"
password = "loral0632"
MD5_input_string = userid+current_date+password
result = hashlib.md5(MD5_input_string.encode())
md5_key = result.hexdigest()
two_yrs_ago_date_value = (datetime.now() - relativedelta(years=2)).strftime('%Y-%m-%d')
print(userid)
print(password)
print(md5_key)
print(current_date)
print(two_yrs_ago_date_value)
print(GQL_Query_list)
gpql_json_data=[]
ticker_value=[]
url = 'http://datastaging1-uat.charles-stanley.co.uk/graphql/'
for i in  GQL_Query_list:
    r = requests.post(url, json={'query': i}, timeout=20)
    response_status_code = r.status_code
    gpql_json_data.append(json.loads(r.text))
print(gpql_json_data)
ticker_value=[]
for j in    gpql_json_data:
    ##gpql_json_data = json.loads(r.text)
    jsonpath_expression = parse('$.data.instrument.nodes[0].symbol')
    symbol = jsonpath_expression.find(j)
    ticker_value.append(symbol[0].value)

print(ticker_value)


url = "http://charlesstanley-intranet.digitallook.com/broker_data_feed?username=charles_stanley_test&key="
ticker = "&ticker="
two_yrs_ago_date = "&date="
url1=[]
for k in ticker_value:
    url1.append(url + md5_key + ticker + k + two_yrs_ago_date + two_yrs_ago_date_value)
print(url1)
xpars=[]
for l in url1:
    data = requests.get(l)
    xpars.append(xmltodict.parse(data.text))
print(xpars)
