import json
from openpyxl import load_workbook
import requests
import xmltodict
import untangle
import hashlib
from jsonpath_ng import jsonpath, parse
wb = load_workbook('C://Users/mkx10/PycharmProjects/cs_datastaging/test_data/LegalEntityPhonenumber_test_data.xlsx')  # Work Book

ws = wb['Sheet1']
xml = ws['A']  # Column
GQL_Query = ws['B']
xml_list = [xml[x].value for x in range (1,(len(xml)))]
GQL_Query_list = [GQL_Query[x].value for x in range (1,(len(GQL_Query)))]

headers = {'content-type': 'text/xml'}

url = "http://csl-jets02.csuat.com:8080/JETS-ci/ciService?wsdl"
SOAP_Response_jSON=[]
for i in  xml_list:
    response = requests.post(url, data=i, headers=headers,timeout=20)
    xpars = xmltodict.parse(response.text)
    SOAP_Response_jSON.append(json.dumps(xpars))




url1 = 'http://datastaging1-sit.charles-stanley.co.uk/graphql/'

gpql_json_data=[]
for j in GQL_Query_list:
    r = requests.post(url1, json={'query': j}, timeout=20)
    gpql_json_data.append(json.loads(r.text))
print(SOAP_Response_jSON)
print(gpql_json_data)
