from openpyxl import load_workbook
import requests

from jsonpath_ng import jsonpath, parse
import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
import warnings

wb = load_workbook('C://cs_datastaging/test_data/LegalEntityRelationshipType_test_data.xlsx')  # Work Book
ws = wb['Sheet1']  # Work Sheet
RelationshipTypeCode = ws['A']  # Column
RelationshipDescription = ws['B']
RelationshipType = ws['C']
ClientAllowed= ws['D']
NonClientAllowed = ws['E']
GQL_Query= ws['F']

RelationshipTypeCode_to_list = [RelationshipTypeCode[x].value for x in range (1,(len(RelationshipTypeCode)))]
RelationshipDescription_to_list = [RelationshipDescription[x].value for x in range (1,(len(RelationshipDescription)))]
RelationshipType_to_list = [RelationshipType[x].value for x in range (1,(len(RelationshipType)))]
ClientAllowed_to_list= [ClientAllowed[x].value for x in range (1,(len(ClientAllowed)))]
NonClientAllowed_to_list = [NonClientAllowed[x].value for x in range (1,(len(NonClientAllowed)))]
GQL_Query_to_list = [GQL_Query[x].value for x in range (1,(len(GQL_Query)))]
print(RelationshipTypeCode_to_list)
print(RelationshipDescription_to_list)
print(RelationshipType_to_list)
print(ClientAllowed_to_list)
print(NonClientAllowed_to_list)
print(GQL_Query_to_list)
gpql_json_data=[]
url = 'http://datastaging1-uat.charles-stanley.co.uk/graphql/'
for i in  GQL_Query_to_list:
    a = requests.post(url, json={'query': i}, timeout=10)
    gpql_json_data.append(json.loads(a.text))
print(gpql_json_data)